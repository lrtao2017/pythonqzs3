#!/usr/bin/env python 
__author__ = "lrtao2010" 

#Create a workbook

#There is no need to create a file on the filesystem to get started with openpyxl.
# Just import the Workbook class and start using it

from openpyxl import Workbook

wb = Workbook()

#A workbook is always created with at least one worksheet.
# You can get it by using the openpyxl.workbook.Workbook.active() property
#This function uses the _active_sheet_index property, set to 0 by default.
# Unless you modify its value, you will always get the first worksheet by using this method.
ws = wb.active

#you can also create new worksheets by using the openpyxl.workbook.
# Workbook.create_sheet() method

ws1 = wb.create_sheet("Mysheet_end") # insert at the end (default)
ws2 = wb.create_sheet("Mysheet_first", 0) # insert at first position

#Sheets are given a name automatically when they are created.
# They are numbered in sequence (Sheet, Sheet1, Sheet2, …).
# You can change this name at any time with the title property:
ws.title = "New Title"

#The background color of the tab holding this title is white by default.
# You can change this providing an RRGGBB color code to
# the sheet_properties.tabColor property:

ws.sheet_properties.tabColor = "FF0000"

#Once you gave a worksheet a name, you can get it as a key of the workbook:

ws3 = wb["New Title"]
print(type(ws3))

#You can review the names of all worksheets
# of the workbook with the openpyxl.workbook.Workbook.sheetnames() property

print(wb.sheetnames)

#You can loop through worksheets
for sheet_name in wb:
    print(sheet_name.title)

#You can create copies of worksheets within a single workbook:
#openpyxl.workbook.Workbook.copy_worksheet() method:

source = wb.active
target = wb.copy_worksheet(source)




#Saving to a file
#The simplest and safest way to save a workbook is
#by using the openpyxl.workbook.Workbook.save() method of
# the openpyxl.workbook.Workbook object:

wb.save("creat_workbook.xlsx")
