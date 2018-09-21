#!/usr/bin/env python 
__author__ = "lrtao2010"

from openpyxl import Workbook

wb = Workbook()

#抓取活动工作表
ws = wb.active

#数据可以直接分配给单元格
ws['A1'] = 'test'

#在一行多列写入数据
ws.append([1,2,3])

#Python类型将自动转换
import datetime
ws['A2'] = datetime.datetime.now()

#在一行多列写入数据
ws.append([4,5,6])

#保存到文件
wb.save("sample.xlsx")


